import io
import re
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────
TREATMENT_HEADERS = [
    "treat_no",
    "treat_lo",
    "treat_name",
    "treat_type",
    "germplasmid",
    "deployid",
    "femaleid",
    "maleid",
    "provenance",
    "species",
    "bri",
    "iri",
]

SPECIES_MAP = {
    "AC": "ACRA",
    "AH": "AHYB",
}

COLUMN_WIDTHS = [10, 10, 80, 14, 14, 14, 14, 14, 14, 10, 8, 8]


# ─────────────────────────────────────────────
# PARSERS
# ─────────────────────────────────────────────
def parse_treatments(text: str) -> list[str]:
    """
    Extract individual treatment names from a Design And Treatment cell.

    Rules (in priority order):
      1. Lines matching 'Tx. ...'  → strip the 'Tx. ' prefix  → keep
      2. Lines starting with '- '  → strip the '- ' prefix    → keep
      3. No matches                → use whole cell text as one treatment

    Subject / header lines (e.g. 'RCBD, 6 Replication',
    'Demo Block Full Compartment', 'Pilot full compartment…') are
    automatically skipped because they don't match either keyword.
    """
    if not text:
        return []

    lines = [l.strip() for l in str(text).splitlines() if l.strip()]

    # ── Priority 1: Tx. pattern ──────────────────────────────────
    tx_lines = [
        re.sub(r'^T\d+\.\s*', '', l).strip()
        for l in lines
        if re.match(r'^T\d+\.', l)
    ]
    if tx_lines:
        return [t for t in tx_lines if t]

    # ── Priority 2: dash bullet pattern ──────────────────────────
    dash_lines = [
        l[1:].strip()           # strip the leading '-'
        for l in lines
        if l.startswith('-')
    ]
    if dash_lines:
        return [d for d in dash_lines if d]

    # ── Fallback: whole cell as a single treatment ────────────────
    return [" ".join(lines)]


def parse_spacing(spacing) -> tuple:
    """
    Extract bri (row inter-row) and iri (intra-row) from a spacing string.

    Examples:
      '3 x 2 m'    → (3, 2)
      '3 x 2.3 m'  → (3, 2.3)
      '3 x 1.66 m' → (3, 1.66)

    Returns integers when the value is whole, floats otherwise.
    """
    if not spacing:
        return None, None

    nums = re.findall(r'[\d.]+', str(spacing))

    def to_num(s):
        v = float(s)
        return int(v) if v == int(v) else v

    bri = to_num(nums[0]) if len(nums) > 0 else None
    iri = to_num(nums[1]) if len(nums) > 1 else None
    return bri, iri


def map_species(raw: str) -> str:
    """Map two-letter species code to full name (AC → ACRA, AH → AHYB)."""
    if not raw:
        return raw
    prefix = str(raw).strip()[:2].upper()
    return SPECIES_MAP.get(prefix, str(raw).strip())


# ─────────────────────────────────────────────
# COLUMN RESOLVER
# ─────────────────────────────────────────────
def resolve_treatment_columns(sheet_headers: dict) -> dict:
    """
    Return a dict of {logical_name: col_index} for all columns
    needed to build the Treatment Plan.
    """
    needed = [
        "Code Trial",
        "Design And Treatment",
        "Title",
        "Species",
        "Spacing",
        "Sector",
        "Comp",
    ]
    return {
        name: sheet_headers[name][0]
        for name in needed
        if name in sheet_headers
    }


# ─────────────────────────────────────────────
# ZIP BUILDER
# ─────────────────────────────────────────────
def build_treatment_zip(
    filtered: list,
    rows_cache: dict,
    treat_cols: dict,
) -> bytes:
    """
    Generate one Treatment Plan Excel per trial and return them
    as a ZIP (folder: Treatment Plan/).

    Args:
        filtered    – list of row-meta dicts from load_database()
        rows_cache  – {row_num: {col: value}} pre-loaded cell cache
        treat_cols  – column index map from resolve_treatment_columns()
    """
    col_design  = treat_cols["Design And Treatment"]
    col_title   = treat_cols["Title"]
    col_species = treat_cols["Species"]
    col_spacing = treat_cols["Spacing"]
    col_code    = treat_cols["Code Trial"]

    # ── Styles (built once, shared across all files) ──────────────
    thin      = Side(style="thin")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell_font = Font(name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_vc   = Alignment(horizontal="left",   vertical="center")
    center    = Alignment(horizontal="center", vertical="center")

    # Columns that look better left-aligned
    LEFT_COLS = {1, 3, 4}     # treat_no, treat_name, treat_type

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for meta in filtered:
            rc = rows_cache[meta["row"]]

            code    = str(rc.get(col_code) or meta["code"]).strip()
            design  = rc.get(col_design)
            title   = rc.get(col_title)
            species = map_species(str(rc.get(col_species) or ""))
            bri, iri = parse_spacing(rc.get(col_spacing))

            # Parse treatments; fallback to Title if Design is empty
            treatments = parse_treatments(design or title)
            if not treatments:
                continue   # nothing to write

            wb = Workbook()
            ws = wb.active
            ws.title = code   # sheet name = Code Trial only

            # ── Header row ───────────────────────────────────────
            for c_idx, hdr in enumerate(TREATMENT_HEADERS, start=1):
                cell           = ws.cell(1, c_idx)
                cell.value     = hdr
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = hdr_align
                cell.border    = border
            ws.row_dimensions[1].height = 28

            # ── Treatment rows ───────────────────────────────────
            for r_idx, treat_name in enumerate(treatments, start=2):
                values = [
                    r_idx - 1,    # treat_no  (continuous, starts at 1)
                    None,         # treat_lo
                    treat_name,   # treat_name
                    "Treatment",  # treat_type  (always)
                    None,         # germplasmid
                    None,         # deployid
                    None,         # femaleid
                    None,         # maleid
                    None,         # provenance
                    species,      # species
                    bri,          # bri
                    iri,          # iri
                ]

                for c_idx, val in enumerate(values, start=1):
                    cell           = ws.cell(r_idx, c_idx)
                    cell.value     = val
                    cell.font      = cell_font
                    cell.border    = border
                    cell.alignment = (
                        left_vc if c_idx in LEFT_COLS else center
                    )

                ws.row_dimensions[r_idx].height = 18

            # ── Column widths ────────────────────────────────────
            for c_idx, width in enumerate(COLUMN_WIDTHS, start=1):
                ws.column_dimensions[
                    ws.cell(1, c_idx).column_letter
                ].width = width

            ws.freeze_panes = "A2"

            out = io.BytesIO()
            wb.save(out)
            zf.writestr(f"Treatment Plan/{meta['filename']}", out.getvalue())

    zip_buf.seek(0)
    return zip_buf.read()
