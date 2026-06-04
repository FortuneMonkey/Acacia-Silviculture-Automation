import io
import datetime
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# ACTIVITY DEFINITIONS
# (activity_code, activity_group) in the exact
# column order they appear in the database.
# ─────────────────────────────────────────────
ACTIVITY_DEFS = [
    ("ASS00",      "Assesment"),
    ("ASS06",      "Assesment"),
    ("ASS12",      "Assesment"),
    ("ASS18",      "Assesment"),
    ("ASS24",      "Assesment"),
    ("ASS30",      "Assesment"),
    ("ASS36",      "Assesment"),
    ("ASS42",      "Assesment"),
    ("ASS48",      "Assesment"),
    ("ASS60",      "Assesment"),
    ("BLK",        "Blanking"),
    ("WR01",       "Weeding Round"),
    ("PDC01",      "Pest & Disease Control"),
    ("WR02",       "Weeding Round"),
    ("FER02",      "Fertilization"),
    ("WR03",       "Weeding Round"),
    ("WR04",       "Weeding Round"),
    ("WR05",       "Weeding Round"),
    ("WR06",       "Weeding Round"),
    ("WR07",       "Weeding Round"),
    ("WR08",       "Weeding Round"),
    ("WR09",       "Weeding Round"),
    ("WR10",       "Weeding Round"),
    ("WR11",       "Weeding Round"),
    ("Singling01", "Singling"),
    ("Singling02", "Singling"),
]

SCHEDULE_HEADERS = [
    "activity_group",
    "activity_code",
    "plan_date",
    "do_date",
    "update_date",
    "required",
    "note",
    "ass_program_by",
    "ass_staff_by",
]

# Afrikaans locale date format  (locale ID 0x0436)
DATE_FMT_AFRIKAANS = "[$-0436]YYYY-MM-DD"

# Per-group row fill colours
GROUP_FILLS = {
    "Assesment":              PatternFill("solid", fgColor="DAEEF3"),
    "Blanking":               PatternFill("solid", fgColor="FDE9D9"),
    "Weeding Round":          PatternFill("solid", fgColor="EBF1DE"),
    "Pest & Disease Control": PatternFill("solid", fgColor="F2DCDB"),
    "Fertilization":          PatternFill("solid", fgColor="FFF2CC"),
    "Singling":               PatternFill("solid", fgColor="E6E0EC"),
}

COLUMN_WIDTHS = [22, 14, 16, 14, 14, 10, 18, 18, 22]


def resolve_activity_columns(sheet_headers: dict) -> dict:
    """
    Map each activity code to its column index in the database.

    Returns:
        act_cols – {activity_code: column_index}
    """
    return {
        code: sheet_headers[code][0]
        for code, _ in ACTIVITY_DEFS
        if code in sheet_headers
    }


def build_schedule_zip(
    filtered: list,
    rows_cache: dict,
    act_cols: dict,
    col_estb: int,
) -> bytes:
    """
    Generate one Schedule Plan Excel per trial and return them
    as a ZIP (folder: Schedule Plan/).

    Args:
        filtered    – list of row-meta dicts from load_database()
        rows_cache  – {row_num: {col: value}} pre-loaded cell cache
        act_cols    – {activity_code: col_index} from resolve_activity_columns()
        col_estb    – column index of 'Established by' in the database
    """
    # ── Styles (built once, shared across all files) ──────────────
    thin      = Side(style="thin")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell_font = Font(name="Arial", size=10)
    center    = Alignment(horizontal="center", vertical="center")
    left_vc   = Alignment(horizontal="left",   vertical="center")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Columns where left-align looks better
    LEFT_ALIGN_COLS = {1, 2, 7, 8, 9}

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for meta in filtered:
            rc       = rows_cache[meta["row"]]
            staff_by = rc.get(col_estb) or ""

            # Collect only activities that have a date value
            act_rows = []
            for code, group in ACTIVITY_DEFS:
                col_idx  = act_cols.get(code)
                if col_idx is None:
                    continue
                date_val = rc.get(col_idx)
                if date_val is None:
                    continue
                # Strip time component — store date only
                if isinstance(date_val, datetime.datetime):
                    date_val = date_val.replace(
                        hour=0, minute=0, second=0, microsecond=0
                    )
                act_rows.append((group, code, date_val))

            if not act_rows:
                continue   # skip trials with no scheduled activities

            wb = Workbook()
            ws = wb.active
            ws.title = meta["filename"].replace(".xlsx", "")

            # ── Header row ────────────────────────────────────────
            for c_idx, hdr in enumerate(SCHEDULE_HEADERS, start=1):
                cell           = ws.cell(1, c_idx)
                cell.value     = hdr
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = hdr_align
                cell.border    = border
            ws.row_dimensions[1].height = 30

            # ── Data rows ─────────────────────────────────────────
            for r_idx, (group, code, date_val) in enumerate(act_rows, start=2):
                row_fill = GROUP_FILLS.get(group, PatternFill())
                values   = [
                    group,     # activity_group
                    code,      # activity_code
                    date_val,  # plan_date
                    None,      # do_date
                    None,      # update_date
                    1,         # required  (always 1)
                    None,      # note
                    None,      # ass_program_by
                    staff_by,  # ass_staff_by  ← Established by
                ]

                for c_idx, val in enumerate(values, start=1):
                    cell           = ws.cell(r_idx, c_idx)
                    cell.value     = val
                    cell.font      = cell_font
                    cell.border    = border
                    cell.fill      = row_fill
                    cell.alignment = (
                        left_vc if c_idx in LEFT_ALIGN_COLS else center
                    )
                    # plan_date → Afrikaans locale date format
                    if c_idx == 3 and isinstance(val, datetime.datetime):
                        cell.number_format = DATE_FMT_AFRIKAANS

                ws.row_dimensions[r_idx].height = 18

            # ── Column widths & freeze ─────────────────────────────
            for c_idx, width in enumerate(COLUMN_WIDTHS, start=1):
                ws.column_dimensions[
                    ws.cell(1, c_idx).column_letter
                ].width = width

            ws.freeze_panes = "A2"

            out = io.BytesIO()
            wb.save(out)
            zf.writestr(f"Schedule Plan/{meta['filename']}", out.getvalue())

    zip_buf.seek(0)
    return zip_buf.read()
