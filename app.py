import io
import datetime
import zipfile
import warnings
warnings.filterwarnings("ignore")

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

from trial_plan     import PLAN_LABELS,       resolve_plan_columns,      build_trial_plan_zip
from schedule_plan  import ACTIVITY_DEFS,     resolve_activity_columns,  build_schedule_zip
from treatment_plan import TREATMENT_HEADERS, resolve_treatment_columns, build_treatment_zip
from treatment_plan import parse_treatments,  parse_spacing

# ═════════════════════════════════════════════
# PAGE CONFIG
# ═════════════════════════════════════════════
st.set_page_config(
    page_title="Acacia's Trial Tools Automation",
    page_icon="🌳",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ═════════════════════════════════════════════
# CUSTOM CSS  (user design, preserved exactly)
# ═════════════════════════════════════════════
st.markdown("""
<style>
    /* ════ Force light theme regardless of system ════ */
    :root { color-scheme: light only; }
    html, body, [data-testid="stAppViewContainer"], .stApp {
        background: #fbfcfb !important;
        color: #1a1a1a !important;
    }
    [data-testid="stHeader"] { background: transparent !important; }
    [data-testid="stSidebar"] {
        background: #ffffff !important;
        border-right: 1px solid #ebefeb !important;
    }

    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, sans-serif !important;
    }

    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 4rem;
        max-width: 1380px;
    }
    #MainMenu, footer, [data-testid="stDecoration"] { visibility: hidden; }

    /* ════ Hero ════ */
    .hero {
        position: relative;
        background: linear-gradient(120deg, #14532d 0%, #166534 45%, #22c55e 130%);
        padding: 2.6rem 2.8rem;
        border-radius: 24px;
        color: white;
        margin-bottom: 2rem;
        overflow: hidden;
        box-shadow: 0 20px 50px -15px rgba(20,83,45,0.45);
    }
    .hero::before {
        content: "";
        position: absolute; top: -60%; right: -8%;
        width: 380px; height: 380px;
        background: radial-gradient(circle, rgba(255,255,255,0.14) 0%, transparent 70%);
        border-radius: 50%;
    }
    .hero::after {
        content: "🌿";
        position: absolute; right: 2.5rem; bottom: -1rem;
        font-size: 7rem; opacity: 0.12; transform: rotate(-12deg);
    }
    .hero h1 {
        font-size: 2.3rem; font-weight: 800; margin: 0;
        letter-spacing: -1px; position: relative; z-index: 1;
    }
    .hero p {
        font-size: 1.05rem; opacity: 0.9; margin: 0.6rem 0 0 0;
        font-weight: 300; max-width: 640px; position: relative; z-index: 1;
    }
    .hero-badge {
        display: inline-block;
        background: rgba(255,255,255,0.18); backdrop-filter: blur(8px);
        border: 1px solid rgba(255,255,255,0.25);
        padding: 0.3rem 0.9rem; border-radius: 30px;
        font-size: 0.78rem; font-weight: 600; letter-spacing: 0.5px;
        margin-bottom: 0.9rem; position: relative; z-index: 1;
    }

    /* ════ Section title ════ */
    .section-title {
        font-size: 1.15rem; font-weight: 700; color: #14532d;
        margin: 1.2rem 0 0.9rem 0;
        display: flex; align-items: center; gap: 0.6rem;
    }
    .section-title .num {
        display: inline-flex; align-items: center; justify-content: center;
        width: 28px; height: 28px; background: #dcfce7; color: #166534;
        border-radius: 8px; font-size: 0.85rem; font-weight: 800;
    }

    /* ════ Metric cards ════ */
    .metric-card {
        background: white; border: 1px solid #edf1ed; border-radius: 16px;
        padding: 1.25rem 1.4rem;
        box-shadow: 0 2px 12px rgba(20,83,45,0.04);
        transition: all 0.2s cubic-bezier(.4,0,.2,1);
        position: relative; overflow: hidden;
    }
    .metric-card::before {
        content: ""; position: absolute; left: 0; top: 0; bottom: 0;
        width: 4px; background: linear-gradient(180deg, #22c55e, #15803d);
    }
    .metric-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 12px 28px rgba(20,83,45,0.12);
        border-color: #c8e6c9;
    }
    .metric-card .icon  { font-size: 1.3rem; margin-bottom: 0.4rem; }
    .metric-card .label {
        font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.8px;
        color: #8a948a; font-weight: 700;
    }
    .metric-card .value {
        font-size: 1.9rem; font-weight: 800; color: #14532d;
        margin-top: 0.15rem; letter-spacing: -0.5px; line-height: 1.1;
    }

    /* ════ Buttons ════ */
    .stButton > button {
        border-radius: 12px; font-weight: 600;
        padding: 0.7rem 1.2rem; border: none;
        transition: all 0.2s ease; font-size: 0.95rem;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(120deg, #166534, #22c55e);
        box-shadow: 0 6px 18px -6px rgba(34,197,94,0.6);
    }
    .stButton > button[kind="primary"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 24px -6px rgba(34,197,94,0.7);
    }
    div[data-testid="stDownloadButton"] > button {
        border-radius: 12px; font-weight: 600;
        background: linear-gradient(120deg, #15803d, #16a34a);
        border: none; box-shadow: 0 6px 18px -6px rgba(22,163,74,0.6);
    }
    div[data-testid="stDownloadButton"] > button:hover { transform: translateY(-2px); }

    /* ════ Tabs ════ */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px; background: #f1f6f1; padding: 7px;
        border-radius: 16px; border: 1px solid #e7eee7;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 11px; padding: 0.6rem 1.5rem; font-weight: 600;
        background: transparent; color: #5a6b5a; transition: all 0.18s ease;
    }
    .stTabs [data-baseweb="tab"]:hover { color: #166534; }
    .stTabs [aria-selected="true"] {
        background: white !important; color: #14532d !important;
        box-shadow: 0 3px 10px rgba(20,83,45,0.1);
    }
    .stTabs [data-baseweb="tab-highlight"],
    .stTabs [data-baseweb="tab-border"] { display: none; }

    /* ════ Tool description card ════ */
    .tool-desc {
        background: linear-gradient(135deg, #ffffff 0%, #f7fbf7 100%);
        border: 1px solid #e4ede4; border-radius: 16px;
        padding: 1.4rem 1.6rem; margin-bottom: 1.4rem;
        transition: all 0.2s ease; height: 100%;
    }
    .tool-desc:hover {
        border-color: #b6e0b6;
        box-shadow: 0 8px 24px rgba(20,83,45,0.07);
    }
    .tool-desc .tool-icon {
        display: inline-flex; align-items: center; justify-content: center;
        width: 46px; height: 46px; background: #dcfce7;
        border-radius: 12px; font-size: 1.4rem; margin-bottom: 0.8rem;
    }
    .tool-desc h4  { margin: 0 0 0.45rem 0; color: #14532d; font-size: 1.1rem; font-weight: 700; }
    .tool-desc p   { margin: 0; color: #5d685d; font-size: 0.92rem; line-height: 1.55; }

    /* ════ Column chips ════ */
    .col-chips { display: flex; flex-wrap: wrap; gap: 6px; margin: 0.4rem 0 1rem 0; }
    .col-chip {
        background: #f1f6f1; border: 1px solid #e0eae0; color: #2d4a2d;
        padding: 0.28rem 0.7rem; border-radius: 8px;
        font-size: 0.78rem; font-family: 'SF Mono','Consolas',monospace; font-weight: 500;
    }

    /* ════ Quality table ════ */
    .quality-ok {
        background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 12px;
        padding: 0.9rem 1.2rem; color: #14532d; font-weight: 600; font-size: 0.95rem;
    }

    /* ════ Generate All card ════ */
    .gen-all-card {
        background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%);
        border: 1.5px solid #86efac; border-radius: 20px;
        padding: 1.8rem 2rem; margin-bottom: 1.5rem;
    }
    .gen-all-card h3 {
        margin: 0 0 0.4rem 0; color: #14532d; font-size: 1.2rem; font-weight: 800;
    }
    .gen-all-card p {
        margin: 0 0 1.2rem 0; color: #4a6b4a; font-size: 0.92rem; line-height: 1.5;
    }

    /* ════ Summary table ════ */
    .summary-banner {
        background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 12px;
        padding: 0.75rem 1.2rem; margin-bottom: 0.8rem;
        display: flex; align-items: center; gap: 0.7rem;
    }
    .summary-banner span { color: #14532d; font-weight: 600; font-size: 0.95rem; }

    /* ════ Alerts ════ */
    [data-testid="stAlert"] { border-radius: 12px; border: none; }

    /* ════ Dataframe ════ */
    [data-testid="stDataFrame"] { border-radius: 14px; overflow: hidden; border: 1px solid #ebefeb; }

    /* ════ Expander ════ */
    [data-testid="stExpander"] {
        border: 1px solid #ebefeb !important; border-radius: 14px !important; background: white;
    }

    /* ════ Sidebar ════ */
    [data-testid="stSidebar"] .sidebar-brand {
        font-size: 1.3rem; font-weight: 800; color: #14532d; margin-bottom: 0.1rem;
    }

    /* ════ Footer ════ */
    .footer {
        text-align: center; color: #a3aaa3; font-size: 0.82rem;
        padding: 2rem 0 0.5rem 0; border-top: 1px solid #eef2ee; margin-top: 2.5rem;
    }
    .footer b { color: #166534; }
</style>
""", unsafe_allow_html=True)


# ═════════════════════════════════════════════
# SESSION STATE  — persists ZIPs across reruns
# ═════════════════════════════════════════════
for _k in ["zip_trial", "zip_sched", "zip_treat", "zip_all",
           "gen_ts", "gen_yr", "summary_df"]:
    if _k not in st.session_state:
        st.session_state[_k] = None


# ═════════════════════════════════════════════
# CONSTANTS
# ═════════════════════════════════════════════
NOT_YET    = {"not yet register", "not yet registered"}
HEADER_ROW = 3
REQUIRED_COLS = ["Code Trial", "Sector", "Comp", "Year", "Register", "Established by"]
ACT_CODES = [c for c, _ in ACTIVITY_DEFS]


# ═════════════════════════════════════════════
# DATABASE LOADER
# ═════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def load_database(file_bytes: bytes):
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    if "Trial Register" not in wb.sheetnames:
        return None, None, None, None, None, None, "Sheet 'Trial Register' not found."

    ws = wb["Trial Register"]
    sheet_headers: dict[str, list[int]] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(HEADER_ROW, col).value
        if raw is not None:
            name = str(raw).strip()
            sheet_headers.setdefault(name, []).append(col)

    missing_req = [c for c in REQUIRED_COLS if c not in sheet_headers]
    if missing_req:
        return None, None, None, None, None, None, (
            f"Required column(s) not found in row {HEADER_ROW}: "
            + ", ".join(f"'{c}'" for c in missing_req)
        )

    plan_cols,  _ = resolve_plan_columns(sheet_headers)
    act_cols      = resolve_activity_columns(sheet_headers)
    treat_cols    = resolve_treatment_columns(sheet_headers)
    special       = {name: sheet_headers[name][0] for name in REQUIRED_COLS}

    needed = (
        set(c for c in plan_cols if c is not None)
        | set(act_cols.values())
        | set(treat_cols.values())
        | set(special.values())
    )

    rows_meta:  list[dict]      = []
    rows_cache: dict[int, dict] = {}

    col_code = special["Code Trial"]
    col_sect = special["Sector"]
    col_comp = special["Comp"]
    col_year = special["Year"]
    col_reg  = special["Register"]

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        code = ws.cell(row, col_code).value
        if not code:
            continue
        reg    = str(ws.cell(row, col_reg).value  or "").strip()
        year   = ws.cell(row, col_year).value
        sector = str(ws.cell(row, col_sect).value or "").strip()
        comp   = str(ws.cell(row, col_comp).value or "").strip()
        rows_meta.append({
            "row": row, "code": str(code).strip(), "register": reg,
            "year": year, "sector": sector, "comp": comp,
            "filename": f"{str(code).strip()} ({sector}{comp}).xlsx",
        })
        if reg.lower() in NOT_YET:
            rows_cache[row] = {col: ws.cell(row, col).value for col in needed}

    return rows_meta, rows_cache, plan_cols, act_cols, treat_cols, special, None


def filter_rows(rows_meta: list, year_filter: str) -> list:
    out = [r for r in rows_meta if r["register"].lower() in NOT_YET]
    if year_filter != "All":
        out = [r for r in out if str(r["year"]) == year_filter]
    return out


# ═════════════════════════════════════════════
# DATA QUALITY CHECK
# ═════════════════════════════════════════════
def data_quality_check(filtered: list, rows_cache: dict, treat_cols: dict,
                        act_cols: dict, col_estb: int) -> pd.DataFrame:
    """
    Scan every filtered trial for common data issues.
    Returns a DataFrame of flagged rows (empty = all clear).
    col_estb is passed directly from special["Established by"] — it is NOT
    inside treat_cols (which only holds Treatment Plan columns).
    """
    col_spacing = treat_cols.get("Spacing")
    col_design  = treat_cols.get("Design And Treatment")
    col_title   = treat_cols.get("Title")

    flags = []
    for meta in filtered:
        rc      = rows_cache[meta["row"]]
        issues  = []

        spacing  = rc.get(col_spacing) if col_spacing is not None else None
        bri, iri = parse_spacing(spacing)
        design   = rc.get(col_design)  if col_design  is not None else None
        title    = rc.get(col_title)   if col_title   is not None else None
        estb     = rc.get(col_estb)    if col_estb    is not None else None

        acts     = sum(1 for c in act_cols.values() if rc.get(c))
        treats   = parse_treatments(design or title)

        if not spacing:
            issues.append("Missing Spacing")
        elif bri is None or iri is None:
            issues.append(f"Unparseable Spacing: '{spacing}'")

        if estb is None or str(estb).strip() == "":
            issues.append("Missing Established By")
        if not design and title:
            issues.append("Design empty — using Title")
        if not treats:
            issues.append("No treatments (file will be skipped)")
        if acts == 0:
            issues.append("No activities (file will be skipped)")

        if issues:
            flags.append({
                "Code Trial":  meta["code"],
                "Sector/Comp": f"{meta['sector']}{meta['comp']}",
                "Year":        meta["year"],
                "Treatments":  len(treats),
                "Activities":  acts,
                "Issues":      " · ".join(issues),
            })

    return pd.DataFrame(flags)


# ═════════════════════════════════════════════
# GENERATE ALL (combined ZIP)
# ═════════════════════════════════════════════
def build_all_plans_zip(filtered, rows_cache, plan_cols,
                        act_cols, treat_cols, col_estb) -> tuple[bytes, pd.DataFrame]:
    """
    Build all three plan types and merge into one ZIP.
    Returns (zip_bytes, summary_df).
    """
    # Build each type
    tp_bytes = build_trial_plan_zip(filtered, rows_cache, plan_cols)
    sc_bytes = build_schedule_zip(filtered, rows_cache, act_cols, col_estb)
    tr_bytes = build_treatment_zip(filtered, rows_cache, treat_cols)

    # Merge into one ZIP
    merged = io.BytesIO()
    with zipfile.ZipFile(merged, "w", zipfile.ZIP_DEFLATED) as zf_out:
        for src_bytes in (tp_bytes, sc_bytes, tr_bytes):
            with zipfile.ZipFile(io.BytesIO(src_bytes)) as zf_in:
                for name in zf_in.namelist():
                    zf_out.writestr(name, zf_in.read(name))

    # Build summary
    col_design  = treat_cols.get("Design And Treatment")
    col_title   = treat_cols.get("Title")
    col_spacing = treat_cols.get("Spacing")

    rows = []
    for meta in filtered:
        rc      = rows_cache[meta["row"]]
        design  = rc.get(col_design)  if col_design  else None
        title   = rc.get(col_title)   if col_title   else None
        spacing = rc.get(col_spacing) if col_spacing else None
        treats  = parse_treatments(design or title)
        acts    = sum(1 for c in act_cols.values() if rc.get(c))
        bri, iri = parse_spacing(spacing)
        rows.append({
            "Code Trial":  meta["code"],
            "Sector/Comp": f"{meta['sector']}{meta['comp']}",
            "Year":        meta["year"],
            "Treatments":  len(treats),
            "Activities":  acts,
            "bri":         bri if bri is not None else "—",
            "iri":         iri if iri is not None else "—",
            "Trial Plan":  "✅",
            "Schedule":    "✅" if acts  > 0 else "⏭ skipped",
            "Treatment":   "✅" if treats else "⏭ skipped",
        })

    merged.seek(0)
    return merged.read(), pd.DataFrame(rows)


# ═════════════════════════════════════════════
# HELPERS
# ═════════════════════════════════════════════
def metric_card(col, icon, label, value):
    col.markdown(
        f'<div class="metric-card"><div class="icon">{icon}</div>'
        f'<div class="label">{label}</div><div class="value">{value}</div></div>',
        unsafe_allow_html=True,
    )

def col_chips(cols: list) -> str:
    chips = "".join(f'<span class="col-chip">{c}</span>' for c in cols)
    return f'<div class="col-chips">{chips}</div>'


# ═════════════════════════════════════════════
# SIDEBAR
# ═════════════════════════════════════════════
with st.sidebar:
    st.markdown('<div class="sidebar-brand">🌳 Trial Tools</div>', unsafe_allow_html=True)
    st.caption("Acacia Silviculture R&D")
    st.divider()

    st.markdown("#####  Upload Database")
    uploaded = st.file_uploader(
        "Excel file — sheet **Trial Register**, headers on **row 3**",
        type=["xlsx"], label_visibility="collapsed",
    )

    st.divider()
    st.markdown("##### ℹ️ How it works")
    st.caption(
        "Upload your **Trial Register** database. The app reads records "
        "marked **'Not yet Register'**, lets you filter by year, then "
        "generates **Trial, Schedule & Treatment Plan** files automatically — "
        "one Excel per trial, bundled into a ZIP."
    )


# ═════════════════════════════════════════════
# HERO
# ═════════════════════════════════════════════
st.markdown("""
<div class="hero">
    <span class="hero-badge">🌱 ACACIA SILVICULTURE R&D TOOLKIT</span>
    <h1>Trial Tools Automation</h1>
    <p>Transform your Trial Register database into ready-to-use Trial, Schedule &amp;
    Treatment Plan files — instantly, and beautifully.</p>
</div>
""", unsafe_allow_html=True)


# ═════════════════════════════════════════════
# NO FILE STATE
# ═════════════════════════════════════════════
if not uploaded:
    c1, c2, c3 = st.columns(3, gap="medium")
    for col, icon, title, desc in [
        (c1, "📋", "Trial Plan",
         "One Excel per trial with all plan fields mapped directly from your database columns."),
        (c2, "📅", "Schedule Plan",
         "One Excel per trial listing every activity that has a scheduled date, neatly structured."),
        (c3, "🧪", "Treatment Plan",
         "One Excel per trial with a parsed row per treatment, extracted from the design notes."),
    ]:
        col.markdown(
            f'<div class="tool-desc"><div class="tool-icon">{icon}</div>'
            f'<h4>{title}</h4><p>{desc}</p></div>',
            unsafe_allow_html=True,
        )
    st.info("👈  **Upload your database file** in the sidebar to begin.")
    st.markdown(
        '<div class="footer">Trial Tools Automation &nbsp;•&nbsp; '
        '<b>Acacia Silviculture R&D 🌳</b></div>',
        unsafe_allow_html=True,
    )
    st.stop()


# ═════════════════════════════════════════════
# LOAD DATABASE
# ═════════════════════════════════════════════
file_bytes = uploaded.read()

with st.spinner("Reading database…"):
    rows_meta, rows_cache, plan_cols, act_cols, treat_cols, special, err = \
        load_database(file_bytes)

if err:
    st.error(f"❌  {err}")
    st.stop()

col_estb = special["Established by"]
st.success(f"✅  Database loaded — **{len(rows_meta)}** total trial record(s) detected.")


# ═════════════════════════════════════════════
# ① FILTER + METRICS
# ═════════════════════════════════════════════
all_not_yet = filter_rows(rows_meta, "All")
available_years = sorted({str(r["year"]) for r in all_not_yet if r["year"]}, reverse=True)

st.markdown(
    '<div class="section-title"><span class="num">1</span> Configure &amp; Preview</div>',
    unsafe_allow_html=True,
)

fcol, _ = st.columns([1, 3])
with fcol:
    year_filter = st.selectbox("📅 Filter by Year", options=["All"] + available_years)

filtered = filter_rows(rows_meta, year_filter)

m1, m2, m3, m4 = st.columns(4, gap="medium")
metric_card(m1, "📊", "Total Records",    len(rows_meta))
metric_card(m2, "⏳", "Not Yet Register", len(all_not_yet))
metric_card(m3, "✅", "Matched (Filter)", len(filtered))
metric_card(m4, "📅", "Active Year",      year_filter)

st.write("")

if not filtered:
    st.warning("No records found with **'Not yet Register'** status for the selected year.")
    st.stop()

with st.expander(f"📑  Preview matched trials  ·  {len(filtered)} records", expanded=True):
    st.dataframe(
        pd.DataFrame([{
            "File Name":  r["filename"], "Code Trial": r["code"],
            "Sector": r["sector"], "Comp": r["comp"], "Year": r["year"],
        } for r in filtered]),
        use_container_width=True, hide_index=True,
        height=min(440, 60 + 36 * len(filtered)),
    )


# ═════════════════════════════════════════════
# ② DATA QUALITY CHECK
# ═════════════════════════════════════════════
st.markdown(
    '<div class="section-title"><span class="num">2</span> Data Quality Check</div>',
    unsafe_allow_html=True,
)

quality_df = data_quality_check(filtered, rows_cache, treat_cols, act_cols, col_estb)

if quality_df.empty:
    st.markdown(
        '<div class="quality-ok">✅  All matched trials passed data quality checks — '
        'no missing fields or skipped files expected.</div>',
        unsafe_allow_html=True,
    )
else:
    n_skip_sched  = quality_df["Issues"].str.contains("No activities").sum()
    n_skip_treat  = quality_df["Issues"].str.contains("No treatments").sum()
    n_title_fb    = quality_df["Issues"].str.contains("using Title").sum()
    n_spacing     = quality_df["Issues"].str.contains("Spacing").sum()
    n_estb        = quality_df["Issues"].str.contains("Established").sum()

    qa1, qa2, qa3, qa4, qa5 = st.columns(5, gap="small")
    metric_card(qa1, "⚠️",  "Flagged Trials",      len(quality_df))
    metric_card(qa2, "⏭",  "Sched. Skipped",       n_skip_sched)
    metric_card(qa3, "⏭",  "Treat. Skipped",       n_skip_treat)
    metric_card(qa4, "📝",  "Title Fallback",        n_title_fb)
    metric_card(qa5, "📐",  "Spacing Issues",        n_spacing)

    st.write("")
    with st.expander(f"🔍  View {len(quality_df)} flagged trial(s)", expanded=False):
        st.dataframe(quality_df, use_container_width=True, hide_index=True)
        st.caption(
            "**Missing Spacing** → `bri`/`iri` will be empty in Treatment Plan.  "
            "**Design empty → using Title** → Treatment name taken from Title column.  "
            "**No activities / No treatments** → that trial's file is skipped for that plan type."
        )


# ═════════════════════════════════════════════
# ③ GENERATE FILES
# ═════════════════════════════════════════════
st.markdown(
    '<div class="section-title"><span class="num">3</span> Generate Files</div>',
    unsafe_allow_html=True,
)

# ── Generate All card ────────────────────────
st.markdown("""
<div class="gen-all-card">
    <h3>⚡ Generate All Plans at Once</h3>
    <p>Builds Trial Plan, Schedule Plan, and Treatment Plan in one go and bundles
    everything into a single ZIP with three sub-folders — no need to click each tab separately.</p>
</div>
""", unsafe_allow_html=True)

ga_col, _ = st.columns([1, 2])
with ga_col:
    gen_all = st.button(
        f"⚡  Generate All  ({len(filtered)} trials × 3 plans)",
        type="primary", use_container_width=True, key="gen_all",
    )

if gen_all:
    prog   = st.progress(0, text="Generating Trial Plans…")
    zip_tp = build_trial_plan_zip(filtered, rows_cache, plan_cols)
    prog.progress(33, text="Generating Schedule Plans…")
    zip_sc = build_schedule_zip(filtered, rows_cache, act_cols, col_estb)
    prog.progress(66, text="Generating Treatment Plans…")
    zip_tr = build_treatment_zip(filtered, rows_cache, treat_cols)
    prog.progress(90, text="Merging into single ZIP…")

    merged = io.BytesIO()
    with zipfile.ZipFile(merged, "w", zipfile.ZIP_DEFLATED) as zf_out:
        for src in (zip_tp, zip_sc, zip_tr):
            with zipfile.ZipFile(io.BytesIO(src)) as zf_in:
                for name in zf_in.namelist():
                    zf_out.writestr(name, zf_in.read(name))
    merged.seek(0)
    st.session_state.zip_all = merged.read()

    # Build summary DataFrame
    col_design  = treat_cols.get("Design And Treatment")
    col_title   = treat_cols.get("Title")
    col_spacing = treat_cols.get("Spacing")
    summary_rows = []
    for meta in filtered:
        rc      = rows_cache[meta["row"]]
        treats  = parse_treatments(rc.get(col_design) or rc.get(col_title))
        acts    = sum(1 for c in act_cols.values() if rc.get(c))
        bri, iri = parse_spacing(rc.get(col_spacing))
        summary_rows.append({
            "Code Trial":  meta["code"],
            "Year":        meta["year"],
            "Treatments":  len(treats),
            "Activities":  acts,
            "bri":         bri if bri is not None else "—",
            "iri":         iri if iri is not None else "—",
            "Trial Plan":  "✅",
            "Schedule":    "✅" if acts  > 0 else "⏭ skipped",
            "Treatment":   "✅" if treats else "⏭ skipped",
        })
    st.session_state.summary_df = pd.DataFrame(summary_rows)

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    yr = year_filter if year_filter != "All" else "ALL"
    st.session_state.gen_ts = ts
    st.session_state.gen_yr = yr

    prog.progress(100, text="Done!")
    st.success(f"✅  All plans generated — {len(filtered)} trials × 3 plan types.")

# Persistent download (survives reruns via session_state)
if st.session_state.zip_all:
    st.download_button(
        label=f"📥  Download All Plans (ZIP)  ·  {len(filtered)} trials × 3 folders",
        data=st.session_state.zip_all,
        file_name=f"All_Plans_{st.session_state.gen_yr}_{st.session_state.gen_ts}.zip",
        mime="application/zip",
        type="primary", use_container_width=True, key="dl_all",
    )
    st.caption(
        "📦  ZIP structure: &nbsp; `Trial Plan/`  &nbsp;·&nbsp;  "
        "`Schedule Plan/`  &nbsp;·&nbsp;  `Treatment Plan/`  "
        "— one Excel per trial in each folder."
    )

# ── Post-generation summary ──────────────────
if st.session_state.summary_df is not None:
    with st.expander("📊  Generation Summary", expanded=True):
        df = st.session_state.summary_df
        c1, c2, c3 = st.columns(3, gap="small")
        c1.metric("Trial Plans",    f"{(df['Trial Plan']=='✅').sum()} ✅")
        c2.metric("Schedule Plans", f"{(df['Schedule']=='✅').sum()} ✅  "
                  f"· {(df['Schedule']!='✅').sum()} skipped")
        c3.metric("Treatment Plans",f"{(df['Treatment']=='✅').sum()} ✅  "
                  f"· {(df['Treatment']!='✅').sum()} skipped")
        st.write("")
        st.dataframe(df, use_container_width=True, hide_index=True,
                     height=min(440, 60 + 36 * len(df)))

st.divider()

# ─────────────────────────────────────────────
# INDIVIDUAL TABS  (session_state download persistence)
# ─────────────────────────────────────────────
st.markdown(
    '<div class="section-title"><span class="num">4</span> '
    'Generate Individual Plans</div>',
    unsafe_allow_html=True,
)
st.caption("Use these tabs if you need only one plan type at a time.")

tab_trial, tab_sched, tab_treat = st.tabs([
    "📋  Trial Plan", "📅  Schedule Plan", "🧪  Treatment Plan",
])

ts_now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
yr_now = year_filter if year_filter != "All" else "ALL"


# ── TAB 1 : TRIAL PLAN ───────────────────────
with tab_trial:
    st.markdown("""
    <div class="tool-desc">
        <div class="tool-icon">📋</div>
        <h4>Trial Plan</h4>
        <p>Generates one Excel file per trial, with all plan fields mapped
        automatically from your database.</p>
    </div>""", unsafe_allow_html=True)

    with st.expander("🔍  View column mapping"):
        st.dataframe(
            pd.DataFrame([{
                "Trial Plan Label": label,
                "Resolved Column":  f"Col {col}" if col else "⚠️ Not found",
            } for label, col in zip(PLAN_LABELS, plan_cols)]),
            use_container_width=True, hide_index=True,
        )

    if st.button(f"⚡  Generate {len(filtered)} Trial Plan(s)",
                 type="primary", key="gen_trial", use_container_width=True):
        with st.spinner(f"Building {len(filtered)} files…"):
            st.session_state.zip_trial = build_trial_plan_zip(filtered, rows_cache, plan_cols)
            st.session_state.gen_ts    = ts_now
            st.session_state.gen_yr    = yr_now
        st.success(f"✅  {len(filtered)} Trial Plan(s) generated.")

    if st.session_state.zip_trial:
        st.download_button(
            "📥  Download Trial Plans (ZIP)", data=st.session_state.zip_trial,
            file_name=f"Trial_Plans_{st.session_state.gen_yr}_{st.session_state.gen_ts}.zip",
            mime="application/zip", type="primary",
            use_container_width=True, key="dl_trial",
        )
        st.caption("📦  ZIP → `Trial Plan/` folder · one Excel per trial.")


# ── TAB 2 : SCHEDULE PLAN ────────────────────
with tab_sched:
    st.markdown("""
    <div class="tool-desc">
        <div class="tool-icon">📅</div>
        <h4>Schedule Plan</h4>
        <p>Generates one Excel per trial listing only activities that have
        a scheduled date.</p>
    </div>""", unsafe_allow_html=True)

    st.markdown("**Output columns**", unsafe_allow_html=True)
    st.markdown(col_chips([
        "activity_group", "activity_code", "plan_date", "do_date",
        "update_date", "required", "note", "ass_program_by", "ass_staff_by",
    ]), unsafe_allow_html=True)

    with st.expander("🔍  View activity column mapping"):
        st.dataframe(
            pd.DataFrame([{
                "Activity Code": code, "Activity Group": group,
                "DB Column": f"Col {act_cols[code]}" if code in act_cols
                             else "⚠️ Not found",
            } for code, group in ACTIVITY_DEFS]),
            use_container_width=True, hide_index=True,
        )
        st.caption(
            f"`required` = 1 always · `ass_staff_by` ← **Established by** "
            f"(Col {col_estb}) · Dates: `[$-0436]YYYY-MM-DD`"
        )

    if st.button(f"⚡  Generate {len(filtered)} Schedule Plan(s)",
                 type="primary", key="gen_sched", use_container_width=True):
        with st.spinner(f"Building {len(filtered)} files…"):
            st.session_state.zip_sched = build_schedule_zip(
                filtered, rows_cache, act_cols, col_estb)
            st.session_state.gen_ts = ts_now
            st.session_state.gen_yr = yr_now
        st.success(f"✅  {len(filtered)} Schedule Plan(s) generated.")

    if st.session_state.zip_sched:
        st.download_button(
            "📥  Download Schedule Plans (ZIP)", data=st.session_state.zip_sched,
            file_name=f"Schedule_Plans_{st.session_state.gen_yr}_{st.session_state.gen_ts}.zip",
            mime="application/zip", type="primary",
            use_container_width=True, key="dl_sched",
        )
        st.caption("📦  ZIP → `Schedule Plan/` folder · one Excel per trial.")


# ── TAB 3 : TREATMENT PLAN ───────────────────
with tab_treat:
    st.markdown("""
    <div class="tool-desc">
        <div class="tool-icon">🧪</div>
        <h4>Treatment Plan</h4>
        <p>Generates one Excel per trial with a row per treatment, parsed from
        <b>Design And Treatment</b> (falls back to <b>Title</b> if empty).</p>
    </div>""", unsafe_allow_html=True)

    st.markdown("**Output columns**", unsafe_allow_html=True)
    st.markdown(col_chips([
        "treat_no","treat_lo","treat_name","treat_type","germplasmid",
        "deployid","femaleid","maleid","provenance","species","bri","iri",
    ]), unsafe_allow_html=True)

    with st.expander("🔍  View parsing rules"):
        st.markdown("""
| Pattern | Action |
|---|---|
| Lines starting with `T1.` `T2.` … | Extracted as individual treatments (prefix stripped) |
| Lines starting with `-` | Extracted as individual treatments (dash stripped) |
| Neither pattern found | Entire cell text used as one treatment |
| Design And Treatment is empty | Falls back to **Title** column |

**Fixed values:** `treat_type` = `Treatment`
**Species mapping:** `AC` → `ACRA` · `AH` → `AHYB`
**Spacing:** `3 x 2.3 m` → `bri = 3`, `iri = 2.3`
        """)

    if st.button(f"⚡  Generate {len(filtered)} Treatment Plan(s)",
                 type="primary", key="gen_treat", use_container_width=True):
        with st.spinner(f"Building {len(filtered)} files…"):
            st.session_state.zip_treat = build_treatment_zip(filtered, rows_cache, treat_cols)
            st.session_state.gen_ts = ts_now
            st.session_state.gen_yr = yr_now
        st.success(f"✅  {len(filtered)} Treatment Plan(s) generated.")

    if st.session_state.zip_treat:
        st.download_button(
            "📥  Download Treatment Plans (ZIP)", data=st.session_state.zip_treat,
            file_name=f"Treatment_Plans_{st.session_state.gen_yr}_{st.session_state.gen_ts}.zip",
            mime="application/zip", type="primary",
            use_container_width=True, key="dl_treat",
        )
        st.caption("📦  ZIP → `Treatment Plan/` folder · one Excel per trial.")


# ═════════════════════════════════════════════
# FOOTER
# ═════════════════════════════════════════════
st.markdown("""
<div class="footer">
    Trial Tools Automation &nbsp;•&nbsp; <b>Acacia Silviculture R&D 🌳</b>
</div>
""", unsafe_allow_html=True)
