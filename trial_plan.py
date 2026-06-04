import io
import datetime
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# FIELD DEFINITIONS
# Labels exactly as they appear in the Trial Plan
# and as headers in the database row 3.
# Duplicates like "SMU" are resolved in order
# of appearance.
# ─────────────────────────────────────────────
PLAN_LABELS = [
    "Code Trial\nNew",
    "Code Trial",
    "Species",
    "Sector",
    "Comp",
    "Secomp",
    "Comp R&D",
    "Material",
    "MSC/MLP",
    "Rotasi",
    "Peat\nDepth (m)",
    "SMU",               # 1st occurrence
    "Site Allocation",
    "Peat Management Zone",
    "theme",
    "group",
    "Title",
    "Serial Trial",
    "Design And Treatment",
    "Planted",
    "Establis",
    "Area  (Ha)",
    "SMU",               # 2nd occurrence
    "Spacing",
    "Peat",
    "Plot Size",
    "status",
    "Close year (note)",
    "Established by",
    "PIC_2022",
    "PIC_2023",
    "PIC_2025",
    "PIC_2026",
    "Assess by",
    "month",
    "Year",
    "remark",
    "Update in STIGMA",
    "PIC stigma",
    "note",
]


def resolve_plan_columns(sheet_headers: dict) -> tuple[list, list]:
    """
    Map each PLAN_LABELS entry to its column index in the database.
    Handles duplicate header names (e.g. 'SMU') by occurrence order.

    Returns:
        plan_cols  – list of column indices (None if label not found)
        missing    – list of label strings that could not be resolved
    """
    label_occ = {}
    plan_cols = []
    missing   = []

    for label in PLAN_LABELS:
        occ       = label_occ.get(label, 0)
        positions = sheet_headers.get(label, [])
        if occ < len(positions):
            plan_cols.append(positions[occ])
        else:
            plan_cols.append(None)
            missing.append(f"'{label}' (occurrence {occ + 1})")
        label_occ[label] = occ + 1

    return plan_cols, missing


def build_trial_plan_zip(filtered: list, rows_cache: dict, plan_cols: list) -> bytes:
    """
    Generate one Trial Plan Excel per trial and return them
    as a ZIP (folder: Trial Plan/).

    Args:
        filtered    – list of row-meta dicts from load_database()
        rows_cache  – {row_num: {col: value}} pre-loaded cell cache
        plan_cols   – column indices resolved by resolve_plan_columns()
    """
    # ── Styles (built once, shared across all files) ──────────────
    thin       = Side(style="thin")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    label_fill = PatternFill("solid", fgColor="D9E1F2")
    title_font = Font(name="Arial", bold=True, size=12, color="1F3864")
    label_font = Font(name="Arial", bold=True, size=10)
    value_font = Font(name="Arial", size=10)
    wrap_top   = Alignment(wrap_text=True, vertical="top")
    left_vc    = Alignment(horizontal="left", vertical="center")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for meta in filtered:
            rc = rows_cache[meta["row"]]

            wb = Workbook()
            ws = wb.active
            ws.title = meta["filename"].replace(".xlsx", "")

            # Title row
            tc           = ws.cell(1, 1)
            tc.value     = f"Trial Plan {meta['code']}"
            tc.font      = title_font
            tc.alignment = left_vc
            ws.merge_cells("A1:B1")
            ws.row_dimensions[1].height = 22

            # Field rows
            for i, (label, col_idx) in enumerate(zip(PLAN_LABELS, plan_cols)):
                r   = i + 2
                val = rc.get(col_idx) if col_idx is not None else None
                if isinstance(val, datetime.datetime):
                    val = val.strftime("%d-%b-%Y")

                lbl           = ws.cell(r, 1)
                lbl.value     = label
                lbl.font      = label_font
                lbl.fill      = label_fill
                lbl.alignment = wrap_top
                lbl.border    = border

                v             = ws.cell(r, 2)
                v.value       = val
                v.font        = value_font
                v.alignment   = wrap_top
                v.border      = border

            ws.column_dimensions["A"].width = 26
            ws.column_dimensions["B"].width = 65

            out = io.BytesIO()
            wb.save(out)
            zf.writestr(f"Trial Plan/{meta['filename']}", out.getvalue())

    zip_buf.seek(0)
    return zip_buf.read()
